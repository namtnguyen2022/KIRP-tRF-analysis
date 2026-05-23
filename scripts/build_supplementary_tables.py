#!/usr/bin/env python3
"""
build_supplementary_tables.py
Creates Supplementary_Tables.xlsx for the KIRP tRF manuscript (MDPI format).
Run with: .venv/bin/python3 scripts/build_supplementary_tables.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

ROOT = "/Users/nam.tnguyen2022/KIRP_tRF_project"
OUT  = os.path.join(ROOT, "Supplementary_Tables.xlsx")

# ── Source files ──────────────────────────────────────────────────────────────
SRC = {
    "S1": os.path.join(ROOT, "results/tables/trf_DE_significant.tsv"),
    "S2": os.path.join(ROOT, "results/tables/gene_DE_full.tsv"),
    "S3": os.path.join(ROOT, "results/tables/anticorrelation_significant.tsv"),
    "S4": os.path.join(ROOT, "data_processed/rna22_binding_results.tsv"),
    "S5": os.path.join(ROOT, "results/tables/roc_auc_top_trfs.tsv"),
    "S6": os.path.join(ROOT, "results/tables/cox_univariate_results.tsv"),
    "S7": os.path.join(ROOT, "results/tables/stage_kruskal_results.tsv"),
}

# ── Style constants ───────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1F3864")   # MDPI dark navy
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
BODY_FONT = Font(name="Calibri", size=10)
NO_WRAP   = Alignment(wrap_text=False, vertical="center")
WRAP_TOP  = Alignment(wrap_text=True,  vertical="top")

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(bottom=THIN)

# Number formats
FMT_SCI   = "0.00E+00"
FMT_4DP   = "0.0000"
FMT_3DP   = "0.000"
FMT_INT   = "0"

# Columns that must render in scientific notation
SCI_COLS = {
    "PValue", "FDR", "p_value", "kruskal_p", "kruskal_fdr",
    "trf_FDR", "gene_FDR",
}

# Tab colours (hex, one per data sheet)
TAB_COLORS = [
    "2E75B6",  # S1 – steel blue
    "2E75B6",  # S2
    "375623",  # S3 – green
    "7030A0",  # S4 – purple
    "C55A11",  # S5 – orange
    "C00000",  # S6 – red
    "404040",  # S7 – dark grey
]

# ── README content ────────────────────────────────────────────────────────────
README_HEADER = ["Table", "Title", "Description", "Rows"]
README_DATA = [
    ("Table S1",
     "Significantly Differentially Expressed tRFs in TCGA-KIRP",
     "All 1,585 tRFs satisfying FDR < 0.05 and |log₂FC| ≥ 1 from edgeR "
     "quasi-likelihood analysis (tumor vs. solid normal). Sorted by FDR.",
     1585),
    ("Table S2",
     "Differential Expression of KIRP Candidate Genes",
     "edgeR quasi-likelihood results for all 16 KIRP-relevant candidate genes "
     "profiled in this study, sorted by FDR.",
     16),
    ("Table S3",
     "Significant Anticorrelated tRF–Gene Pairs",
     "All 104 tRF–gene pairs with Spearman r < 0 and FDR < 0.05 following "
     "the tRF-down/gene-up biological pattern, across 290 matched tumor samples.",
     104),
    ("Table S4",
     "RNA22 v2 Predicted tRF–3′UTR Binding Interactions",
     "All 130 RNA22 v2 binding-site predictions for the 20 most anticorrelated "
     "tRFs against the MET, TERT, and CDKN2A 3′UTRs; includes heteroduplex "
     "structures. Significant column flags p < 0.05 predictions.",
     130),
    ("Table S5",
     "Internal Tumor-Normal Discrimination AUC Results",
     "AUC values from ROC analysis for the top 20 differentially expressed tRFs "
     "(ranked by |log₂FC|), distinguishing KIRP tumor from solid normal "
     "tissue within TCGA-KIRP. Sorted by AUC descending.",
     20),
    ("Table S6",
     "Exploratory Univariable Cox Regression: Overall Survival",
     "Univariable Cox proportional-hazards regression for the top 20 "
     "differentially expressed tRFs using overall survival (251 tumor samples, "
     "44 death events). Log₂(CPM+1) as continuous predictor; no covariate "
     "adjustment. Results are exploratory.",
     20),
    ("Table S7",
     "Kruskal–Wallis tRF-Stage Association Results",
     "Kruskal–Wallis test results for stage-dependent expression across "
     "AJCC pathologic stages I–IV in 261 staged primary tumor samples, "
     "for the six top AUC tRF candidates. FDR by Benjamini–Hochberg.",
     6),
]

# ── Helpers ───────────────────────────────────────────────────────────────────

def col_width(df: pd.DataFrame, col: str, extra: int = 2) -> float:
    """Estimate column width from header + data (capped)."""
    header_len = len(col)
    if col == "heteroduplex":
        return 85
    try:
        data_len = df[col].astype(str).str.len().quantile(0.95)
    except Exception:
        data_len = header_len
    raw = max(header_len, data_len) + extra
    # tRF id columns can be wide; everything else capped at 50
    if any(k in col.lower() for k in ("trf", "trf_id")):
        return max(float(raw) * 1.05, 26)
    return min(float(raw) * 1.05, 50)


def write_header_row(ws, columns):
    for ci, col in enumerate(columns, 1):
        c = ws.cell(1, ci)
        c.value    = col
        c.font     = HDR_FONT
        c.fill     = HDR_FILL
        c.alignment = NO_WRAP
        c.border   = THIN_BORDER


def write_data_rows(ws, df: pd.DataFrame, sci_cols: set):
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, (col, val) in enumerate(zip(df.columns, row), 1):
            c = ws.cell(ri, ci)
            c.font      = BODY_FONT
            c.alignment = NO_WRAP if col != "heteroduplex" else WRAP_TOP

            if pd.isna(val) if isinstance(val, float) else False:
                c.value = ""
            elif col in sci_cols and isinstance(val, float):
                c.value          = val
                c.number_format  = FMT_SCI
            elif isinstance(val, float):
                c.value          = val
                c.number_format  = FMT_4DP
            elif isinstance(val, (int, bool)):
                c.value = val
            else:
                c.value = str(val) if val is not None else ""


def apply_col_widths(ws, df: pd.DataFrame):
    for ci, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_width(df, col)


def make_data_sheet(wb, sheet_name: str, df: pd.DataFrame,
                    sci_cols: set, tab_color: str):
    ws = wb.create_sheet(sheet_name)
    ws.freeze_panes         = "A2"
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines  = True

    write_header_row(ws, df.columns.tolist())
    write_data_rows(ws, df, sci_cols)
    apply_col_widths(ws, df)
    ws.row_dimensions[1].height = 16

    return ws

# ── README sheet ──────────────────────────────────────────────────────────────

def make_readme(wb):
    ws = wb.create_sheet("README", 0)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "000000"

    # Manuscript title
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value     = ("Supplementary Tables — Computational Analysis Identifies "
                   "Candidate tRNA-Derived Fragment–mRNA Regulatory "
                   "Biomarkers in Papillary Renal Cell Carcinoma")
    t.font      = Font(bold=True, size=12, name="Calibri", color="1F3864")
    t.alignment = Alignment(wrap_text=True, horizontal="left", vertical="top")
    ws.row_dimensions[1].height = 36

    # Blank row
    ws.append([])
    ws.row_dimensions[2].height = 6

    # Header
    ws.append(README_HEADER)
    for ci in range(1, 5):
        c = ws.cell(3, ci)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL
        c.alignment = NO_WRAP
    ws.row_dimensions[3].height = 16

    # Data rows
    for i, row in enumerate(README_DATA, 4):
        ws.append(list(row))
        ws.cell(i, 1).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(i, 2).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(i, 4).alignment = Alignment(horizontal="center", vertical="top")
        for ci in range(1, 5):
            c = ws.cell(i, ci)
            if ci != 4:
                c.alignment = WRAP_TOP
            if ci not in (1, 2):
                c.font = Font(name="Calibri", size=10)
        ws.row_dimensions[i].height = 52

    ws.freeze_panes = "A4"

    # Column widths
    for col, width in zip("ABCD", [14, 52, 88, 8]):
        ws.column_dimensions[col].width = width

# ── Data loaders ──────────────────────────────────────────────────────────────

def load_s1():
    df = pd.read_csv(SRC["S1"], sep="\t")
    df = df.sort_values("FDR").reset_index(drop=True)
    return df, {"PValue", "FDR"}


def load_s2():
    df = pd.read_csv(SRC["S2"], sep="\t")
    df = df.sort_values("FDR").reset_index(drop=True)
    return df, {"PValue", "FDR"}


def load_s3():
    df = pd.read_csv(SRC["S3"], sep="\t")
    df = df.sort_values("FDR").reset_index(drop=True)
    return df, {"p_value", "FDR", "trf_FDR", "gene_FDR"}


def load_s4():
    df = pd.read_csv(SRC["S4"], sep="\t")
    # Add significance flag right after p_value column
    p_idx = df.columns.get_loc("p_value") + 1
    df.insert(p_idx, "significant_p05",
              df["p_value"].apply(lambda x: "Yes" if x < 0.05 else "No"))
    df = df.sort_values("p_value").reset_index(drop=True)
    return df, {"p_value"}


def load_s5():
    df = pd.read_csv(SRC["S5"], sep="\t")
    df.columns = ["trf_id", "AUC"]
    df = df.sort_values("AUC", ascending=False).reset_index(drop=True)
    return df, set()


def load_s6():
    df = pd.read_csv(SRC["S6"], sep="\t")
    df = df.rename(columns={
        "tRF"  : "trf_id",
        "HR_lo": "lower_95_CI",
        "HR_hi": "upper_95_CI",
    })
    df = df.sort_values("FDR").reset_index(drop=True)
    return df, {"p_value", "FDR"}


def load_s7():
    df = pd.read_csv(SRC["S7"], sep="\t")
    df = df.rename(columns={
        "tRF"        : "trf_id",
        "kruskal_p"  : "p_value",
        "kruskal_fdr": "FDR",
    })
    df.insert(len(df.columns), "significant_FDR05",
              df["FDR"].apply(lambda x: "Yes" if x < 0.05 else "No"))
    df = df.sort_values("FDR").reset_index(drop=True)
    return df, {"p_value", "FDR"}


SHEETS = [
    ("Table S1", load_s1),
    ("Table S2", load_s2),
    ("Table S3", load_s3),
    ("Table S4", load_s4),
    ("Table S5", load_s5),
    ("Table S6", load_s6),
    ("Table S7", load_s7),
]

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    make_readme(wb)
    print("  ✓ README")

    for (sheet_name, loader), tab_color in zip(SHEETS, TAB_COLORS):
        df, sci_cols = loader()
        make_data_sheet(wb, sheet_name, df, sci_cols, tab_color)
        print(f"  ✓ {sheet_name}: {len(df):,} rows × {len(df.columns)} columns")

    wb.save(OUT)
    print(f"\n✅  Saved → {OUT}")


if __name__ == "__main__":
    main()
